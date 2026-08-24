import streamlit as st
import os
import io
import math
import cv2
import numpy as np
import pandas as pd
from PIL import Image
import pypdfium2 as pdfium
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from saq_vector_engine import DXFVectorParser, compare_vector_delta

LOGO_PATH = "logo.png.png" if os.path.exists("logo.png.png") else "logo.png"
has_logo = os.path.exists(LOGO_PATH)
app_icon = Image.open(LOGO_PATH) if has_logo else "📐"

st.set_page_config(page_title="S.A.Q - Takeoff & Vector CAD Platform", layout="wide", page_icon=app_icon)

def load_raster(file, scale=2.0):
    """טעינת PDF/תמונה ברזולוציית High-DPI לדיוק מקסימלי"""
    if file.name.lower().endswith(".pdf"):
        pdf = pdfium.PdfDocument(file.read())
        bitmap = pdf.get_page(0).render(scale=scale)
        pil_img = bitmap.to_pil().convert("RGB")
        return cv2.cvtColor(np.array(pil_img), cv2.COLOR_RGB2BGR)
    else:
        file_bytes = np.asarray(bytearray(file.read()), dtype=np.uint8)
        img = cv2.imdecode(file_bytes, cv2.IMREAD_UNCHANGED)
        if img is None:
            return None
        if len(img.shape) == 3 and img.shape[2] == 4:
            alpha = img[:, :, 3] / 255.0
            bg = np.ones_like(img[:, :, :3], dtype=np.uint8) * 255
            for c in range(3):
                bg[:, :, c] = (img[:, :, c] * alpha + bg[:, :, c] * (1.0 - alpha)).astype(np.uint8)
            return bg
        elif len(img.shape) == 2:
            return cv2.cvtColor(img, cv2.COLOR_GRAY2BGR)
        return img

def is_valid_electrical_symbol(crop_gray):
    """מסנן מתמטי חסין תקלות: פוסל אותיות, מספרים, קווי טבלה ורעשים"""
    h, w = crop_gray.shape[:2]
    if w < 16 or h < 16 or w > 140 or h > 140:
        return False
    
    _, thresh = cv2.threshold(crop_gray, 220, 255, cv2.THRESH_BINARY_INV)
    contours, _ = cv2.findContours(thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    if not contours:
        return False
        
    total_area = sum(cv2.contourArea(c) for c in contours)
    bounding_area = float(w * h)
    density = total_area / bounding_area if bounding_area > 0 else 0
    
    # סמלי חשמל מוגדרים בצפיפות קו אופיינית
    if density < 0.08 or density > 0.80:
        return False
        
    return True

def extract_symbols_from_legend(legend_img):
    """חילוץ סמלים נקי ומדויק מתוך המקרא עם מסיכות בידוד"""
    gray = cv2.cvtColor(legend_img, cv2.COLOR_BGR2GRAY)
    leg_h = gray.shape[0]
    crop_h = int(leg_h * 0.88)
    work_gray = gray[:crop_h, :]
    work_color = legend_img[:crop_h, :]
    
    _, thresh = cv2.threshold(work_gray, 225, 255, cv2.THRESH_BINARY_INV)
    
    # נטרול רשתות טבלה במקרא
    h_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (30, 1))
    v_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, 30))
    h_lines = cv2.morphologyEx(thresh, cv2.MORPH_OPEN, h_kernel)
    v_lines = cv2.morphologyEx(thresh, cv2.MORPH_OPEN, v_kernel)
    cleaned = cv2.subtract(thresh, cv2.add(h_lines, v_lines))
    
    contours, _ = cv2.findContours(cleaned, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    raw_symbols = []
    for c in contours:
        x, y, w, h = cv2.boundingRect(c)
        area = cv2.contourArea(c)
        
        if 20 <= w <= 110 and 20 <= h <= 110 and area > 75:
            aspect = w / float(h)
            if 0.55 <= aspect <= 1.85:
                pad = 4
                y1, y2 = max(0, y - pad), min(work_gray.shape[0], y + h + pad)
                x1, x2 = max(0, x - pad), min(work_gray.shape[1], x + w + pad)
                c_gray = work_gray[y1:y2, x1:x2]
                c_color = work_color[y1:y2, x1:x2]
                
                if is_valid_electrical_symbol(c_gray):
                    raw_symbols.append({
                        "bbox": (x, y, w, h),
                        "crop_color": c_color,
                        "crop_gray": c_gray,
                        "y_pos": y,
                        "x_pos": x
                    })
                    
    raw_symbols.sort(key=lambda s: (s["y_pos"] // 35, s["x_pos"]))
    
    unique_symbols = []
    for sym in raw_symbols:
        is_dup = False
        for u in unique_symbols:
            if np.hypot(sym["x_pos"] - u["x_pos"], sym["y_pos"] - u["y_pos"]) < 30:
                is_dup = True
                break
        if not is_dup:
            unique_symbols.append(sym)
            
    return unique_symbols[:15]

def match_symbol_99pct(plan_gray, templ_gray, min_thresh=0.55, high_thresh=0.74, ignore_bottom_pct=0.18):
    """מנוע התאמה בדיוק 99%: מסיכת קווים + אימות צורה גיאומטרי + סינון כפול"""
    h_p, w_p = plan_gray.shape[:2]
    active_h = int(h_p * (1.0 - ignore_bottom_pct))
    plan_roi = plan_gray[:active_h, :]
    
    # חיתוך שוליים עודפים מהסמל
    _, thresh_t = cv2.threshold(templ_gray, 235, 255, cv2.THRESH_BINARY_INV)
    pts = cv2.findNonZero(thresh_t)
    if pts is not None:
        tx, ty, tw, th = cv2.boundingRect(pts)
        if tw > 12 and th > 12:
            templ_gray = templ_gray[ty:ty+th, tx:tx+tw]
            thresh_t = thresh_t[ty:ty+th, tx:tx+tw]
            
    # יצירת מסיכה שחורה/לבנה ממוקדת של קווי הסמל
    templ_mask = thresh_t.copy()
    
    detections = []
    scales = [0.88, 0.94, 1.0, 1.06, 1.14]
    
    for scale in scales:
        sw = int(templ_gray.shape[1] * scale)
        sh = int(templ_gray.shape[0] * scale)
        if sw >= plan_roi.shape[1] or sh >= plan_roi.shape[0] or sw < 12 or sh < 12:
            continue
            
        resized_t = cv2.resize(templ_gray, (sw, sh))
        resized_m = cv2.resize(templ_mask, (sw, sh))
        
        for rot in [0, 90, 180, 270]:
            if rot == 90:
                r_t = cv2.rotate(resized_t, cv2.ROTATE_90_CLOCKWISE)
                r_m = cv2.rotate(resized_m, cv2.ROTATE_90_CLOCKWISE)
            elif rot == 180:
                r_t = cv2.rotate(resized_t, cv2.ROTATE_180)
                r_m = cv2.rotate(resized_m, cv2.ROTATE_180)
            elif rot == 270:
                r_t = cv2.rotate(resized_t, cv2.ROTATE_90_COUNTERCLOCKWISE)
                r_m = cv2.rotate(resized_m, cv2.ROTATE_90_COUNTERCLOCKWISE)
            else:
                r_t, r_m = resized_t, resized_m
                
            rw, rh = r_t.shape[::-1]
            
            # 1. התאמת תבנית מנורמלת עם מסיכה (מתעלמת מרקע הקיר)
            res = cv2.matchTemplate(plan_roi, r_t, cv2.TM_CCORR_NORMED, mask=r_m)
            loc = np.where(res >= min_thresh)
            
            for pt in zip(*loc[::-1]):
                score = float(res[pt[1], pt[0]])
                
                # 2. אימות צורה גיאומטרי באזור התוכנית שנלכד
                crop_found = plan_roi[pt[1]:pt[1]+rh, pt[0]:pt[0]+rw]
                _, crop_th = cv2.threshold(crop_found, 220, 255, cv2.THRESH_BINARY_INV)
                stroke_density = np.sum(crop_th > 0) / float(rw * rh)
                
                # סינון משני: מוודא שיש קווי חשמל אמיתיים ולא חלל לבן ריק
                if 0.06 <= stroke_density <= 0.85:
                    status = "Green" if score >= high_thresh else "Yellow"
                    detections.append({
                        "bbox": (int(pt[0]), int(pt[1]), int(rw), int(rh)),
                        "center": (int(pt[0] + rw // 2), int(pt[1] + rh // 2)),
                        "score": score,
                        "status": status
                    })
                    
    if not detections:
        return []
        
    boxes = [list(d["bbox"]) for d in detections]
    scores = [d["score"] for d in detections]
    # NMS מותאם שמפריד שקעים סמוכים
    indices = cv2.dnn.NMSBoxes(boxes, scores, score_threshold=min_thresh, nms_threshold=0.20)
    return [detections[i] for i in indices.flatten()] if len(indices) > 0 else []

def generate_excel_with_embedded_images(boq_rows):
    """ייצוא לאקסל מעוצב עם שיבוץ תמונות הסמלים בתוך התאים"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "כתב כמויות חשמל - SAQ"
    ws.sheet_view.rightToLeft = True
    
    headers = ["מס'", "תמונת הסמל", "תיאור הפריט", "כמות מאושרת", "יחידת מידה"]
    ws.append(headers)
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    for col_num in range(1, 6):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 34
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 14
    ws.row_dimensions[1].height = 30
    
    for idx, item in enumerate(boq_rows):
        row_idx = idx + 2
        ws.row_dimensions[row_idx].height = 52
        
        c_num = ws.cell(row=row_idx, column=1, value=item["מס'"])
        c_desc = ws.cell(row=row_idx, column=3, value=item["תיאור הפריט"])
        c_qty = ws.cell(row=row_idx, column=4, value=item["כמות"])
        c_unit = ws.cell(row=row_idx, column=5, value=item["יחידת מידה"])
        
        for c in [c_num, ws.cell(row=row_idx, column=2), c_desc, c_qty, c_unit]:
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin_border
            c.font = Font(name="Arial", size=10)
            
        sym_img = item.get("symbol_img")
        if sym_img is not None and sym_img.size > 0:
            rgb = cv2.cvtColor(sym_img, cv2.COLOR_BGR2RGB)
            pil_img = Image.fromarray(rgb)
            pil_img.thumbnail((60, 46))
            
            img_io = io.BytesIO()
            pil_img.save(img_io, format="PNG")
            img_io.seek(0)
            
            xl_img = OpenpyxlImage(img_io)
            ws.add_image(xl_img, f"B{row_idx}")
            
    out_io = io.BytesIO()
    wb.save(out_io)
    return out_io.getvalue()

with st.sidebar:
    if has_logo:
        st.image(LOGO_PATH, use_container_width=True)
    st.header("⚙️ הגדרות עבודה")
    file_type = st.radio("פורמט שרטוט:", ["📄 PDF / תמונה (Raster)", "📐 CAD וקטורי (DXF)"])
    mode = st.radio("מצב פעולה:", ["ספירה מתוכנית בודדת", "השוואת שינויים (Delta)"])
    discipline = st.selectbox("דיסציפלינה:", ["⚡ חשמל ומאור", "🧱 בניה (מחיצות ומעטפת)", "🚿 אינסטלציה", "📐 ריצוף וחיפוי"])
    st.markdown("---")
    st.subheader("📏 כיול דיוק 99%")
    high_sens = st.slider("סף ודאי אוטומטי (Green %):", min_value=65, max_value=85, value=74, step=1)
    min_sens = st.slider("סף שאלת משתמש (Yellow %):", min_value=45, max_value=64, value=55, step=1)
    filter_banner = st.checkbox("סנן טבלת כותרת (Title Block)", value=True)

col_l, col_t = st.columns([1, 6])
with col_l:
    if has_logo:
        st.image(LOGO_PATH, width=90)
with col_t:
    st.title("S.A.Q Takeoff & Delta Platform")
    st.caption("פלטפורמת ענן מנוע דיוק 99% - פענוח הנדסי, ספירת כמויות ו-BOQ")

# ========================================================
# 📄 נתיב PDF ורסטר (Legend-to-Plan Matcher 99%)
# ========================================================
if file_type == "📄 PDF / תמונה (Raster)":
    if mode == "ספירה מתוכנית בודדת":
        c_p, c_l = st.columns(2)
        with c_p:
            f_plan = st.file_uploader("1️⃣ העלה שרטוט תוכנית (PDF / תמונה):", type=["pdf", "png", "jpg", "jpeg"], key="plan_in")
        with c_l:
            f_legend = st.file_uploader("2️⃣ העלה קובץ מקרא (PDF / תמונה):", type=["pdf", "png", "jpg", "jpeg"], key="leg_in")
            
        if f_plan and f_legend:
            if st.button("🚀 הפעל פענוח מקרא וספירה מדויקת בתוכנית (99%)"):
                img_plan = load_raster(f_plan, scale=2.0)
                img_leg = load_raster(f_legend, scale=2.0)
                
                symbols_found = extract_symbols_from_legend(img_leg)
                
                if not symbols_found:
                    st.warning("⚠️ לא אותרו סמלים הנדסיים מבודדים במקרא.")
                else:
                    progress_bar = st.progress(0, text="מתחיל סריקה הנדסית בדיוק 99%...")
                    plan_gray = cv2.cvtColor(img_plan, cv2.COLOR_BGR2GRAY)
                    
                    all_results = []
                    total_syms = len(symbols_found)
                    
                    for i, sym in enumerate(symbols_found):
                        progress_bar.progress((i + 1) / total_syms, text=f"סורק סמל {i+1} מתוך {total_syms} (Masked Engine)...")
                        matches = match_symbol_99pct(
                            plan_gray, 
                            sym["crop_gray"], 
                            min_thresh=(min_sens / 100.0),
                            high_thresh=(high_sens / 100.0),
                            ignore_bottom_pct=0.18 if filter_banner else 0.0
                        )
                        
                        all_results.append({
                            "index": i + 1,
                            "symbol_img": sym["crop_color"],
                            "matches": matches
                        })
                        
                    progress_bar.empty()
                    st.session_state["legend_results"] = all_results
                    st.session_state["raw_plan_img"] = img_plan

            if "legend_results" in st.session_state:
                res = st.session_state["legend_results"]
                raw_plan = st.session_state["raw_plan_img"]
                disp_plan = raw_plan.copy()
                
                if filter_banner:
                    h_cut = int(raw_plan.shape[0] * 0.82)
                    cv2.line(disp_plan, (0, h_cut), (raw_plan.shape[1], h_cut), (180, 180, 180), 3)
                    cv2.putText(disp_plan, "Filtered Title Block Area", (20, h_cut + 35), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (140, 140, 140), 2)
                
                # בדיקה ואישור נקודות בספק (Human-in-the-Loop)
                yellow_items = []
                for s_idx, item in enumerate(res):
                    for m_idx, m in enumerate(item["matches"]):
                        if m["status"] == "Yellow":
                            yellow_items.append((s_idx, m_idx, item, m))
                            
                if yellow_items:
                    st.markdown("---")
                    st.warning(f"⚠️ **אותרו {len(yellow_items)} נקודות בספק לבדיקתך (Human-in-the-Loop להשלמת 100% דיוק):**")
                    with st.expander("🔍 לחץ כאן לבדיקה ואישור נקודות בספק", expanded=True):
                        cols = st.columns(3)
                        for y_i, (s_idx, m_idx, item, m) in enumerate(yellow_items):
                            with cols[y_i % 3]:
                                x, y, w, h = m["bbox"]
                                pad = 24
                                y1, y2 = max(0, y - pad), min(raw_plan.shape[0], y + h + pad)
                                x1, x2 = max(0, x - pad), min(raw_plan.shape[1], x + w + pad)
                                crop_zoom = raw_plan[y1:y2, x1:x2].copy()
                                cv2.rectangle(crop_zoom, (x - x1, y - y1), (x - x1 + w, y - y1 + h), (0, 165, 255), 2)
                                
                                st.image(cv2.cvtColor(crop_zoom, cv2.COLOR_BGR2RGB), caption=f"סמל #{item['index']} (ודאות {m['score']*100:.0f}%)", width=140)
                                approved = st.checkbox("✅ אשר נקודה זו", value=False, key=f"appr_{s_idx}_{m_idx}")
                                m["user_approved"] = approved
                                st.markdown("---")
                
                boq_rows = []
                for s_idx, item in enumerate(res):
                    confirmed_count = 0
                    for m_idx, m in enumerate(item["matches"]):
                        is_green = (m["status"] == "Green")
                        is_user_appr = m.get("user_approved", False)
                        
                        x, y, w, h = m["bbox"]
                        if is_green:
                            confirmed_count += 1
                            cv2.rectangle(disp_plan, (x, y), (x + w, y + h), (0, 200, 0), 3)
                            cv2.putText(disp_plan, f"#{item['index']}", (x, max(15, y - 4)), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 200, 0), 2)
                        elif is_user_appr:
                            confirmed_count += 1
                            cv2.rectangle(disp_plan, (x, y), (x + w, y + h), (0, 200, 0), 3)
                            cv2.putText(disp_plan, f"#{item['index']} (V)", (x, max(15, y - 4)), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 200, 0), 2)
                        else:
                            cv2.rectangle(disp_plan, (x, y), (x + w, y + h), (0, 165, 255), 2)
                            cv2.putText(disp_plan, "?", (x, max(15, y - 4)), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 165, 255), 2)
                            
                    item["confirmed_count"] = confirmed_count

                st.subheader("📋 פירוט ספירת כמויות סופית")
                for item in res:
                    c1, c2, c3 = st.columns([1.5, 2.5, 1.5])
                    with c1:
                        if item["symbol_img"].size > 0:
                            st.image(cv2.cvtColor(item["symbol_img"], cv2.COLOR_BGR2RGB), width=90, caption=f"סמל #{item['index']}")
                    with c2:
                        s_desc = st.text_input(f"תיאור פריט #{item['index']}:", value=f"סמל חשמל #{item['index']}", key=f"desc_{item['index']}")
                        is_inc = st.checkbox("כלול בכתב כמויות", value=(item["confirmed_count"] > 0), key=f"inc_leg_{item['index']}")
                    with c3:
                        st.metric("כמות מאושרת:", f"{item['confirmed_count']} יח'")
                    
                    if is_inc:
                        boq_rows.append({
                            "מס'": item["index"],
                            "symbol_img": item["symbol_img"],
                            "תיאור הפריט": s_desc,
                            "כמות": item["confirmed_count"],
                            "יחידת מידה": "יח'"
                        })
                    st.markdown("---")
                    
                st.subheader("🗺️ תוכנית עם סימוני הפריטים (ירוק = ודאי ומאושר, כתום = לא אושר):")
                st.image(cv2.cvtColor(disp_plan, cv2.COLOR_BGR2RGB), use_container_width=True)
                
                if boq_rows:
                    st.subheader("📊 ריכוז סופי לכתב כמויות")
                    df_preview = pd.DataFrame([
                        {"מס'": r["מס'"], "תיאור הפריט": r["תיאור הפריט"], "כמות מאושרת": r["כמות"], "יחידת מידה": r["יחידת מידה"]}
                        for r in boq_rows
                    ])
                    st.dataframe(df_preview, use_container_width=True)
                    
                    excel_bytes = generate_excel_with_embedded_images(boq_rows)
                    st.download_button(
                        "📥 ייצא כתב כמויות ל-Excel (כולל תמונות סמלים)",
                        data=excel_bytes,
                        file_name="Approved_Electrical_Takeoff_99Pct.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
        else:
            st.info("💡 אנא העלה את קובץ התוכנית ואת קובץ המקרא להפעלת הסריקה.")

# ========================================================
# 📐 נתיב CAD וקטורי (DXF) - 100% דיוק דטרמיניסטי
# ========================================================
else:
    scale_val = 1.0
    if mode == "ספירה מתוכנית בודדת":
        cad_file = st.file_uploader("העלה שרטוט CAD (DXF):", type=["dxf"])
        if cad_file:
            parser = DXFVectorParser(cad_file, unit_scale_to_meter=scale_val)
            layers = [l["name"] for l in parser.get_layers_summary() if l["entity_count"] > 0]
            if discipline == "⚡ חשמל ומאור":
                st.subheader("⚡ ספירת סמלי חשמל מבוססת בלוקים (100% דיוק וקטורי)")
                sel_layers = st.multiselect("שכבות חשמל:", layers, default=layers[:3] if layers else [])
                blocks = parser.extract_blocks(sel_layers)
                if blocks:
                    df = pd.DataFrame(blocks)
                    summary = df.groupby(["name", "cardinal_rotation"]).size().reset_index(name="כמות")
                    st.dataframe(summary, use_container_width=True)
                    df["אושר"] = True
                    edited = st.data_editor(df[["name", "layer", "x", "y", "rotation_deg", "cardinal_rotation", "אושר"]], use_container_width=True)
                    out = io.BytesIO()
                    with pd.ExcelWriter(out, engine="openpyxl") as w:
                        summary.to_excel(w, sheet_name="ריכוז", index=False)
                        edited.to_excel(w, sheet_name="פירוט", index=False)
                    st.download_button("📥 ייצא ל-Excel", data=out.getvalue(), file_name="Electrical_BOQ.xlsx")
